from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import csv
import unicodedata
import re

# Importações adicionais para o Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def remove_accents(input_str):
    """
    Remove acentos e caracteres especiais de uma string.
    """
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    only_ascii = ''.join([c for c in nfkd_form if not unicodedata.combining(c)])
    # Remove caracteres especiais, deixando apenas letras, números e espaços
    return re.sub(r'[^a-zA-Z0-9 ]', '', only_ascii)

def parse_report_text(report_text):
    """
    Analisa o texto do relatório e retorna uma estrutura de dados.
    """
    lines = report_text.strip().split('\n')
    notas = {'resumo': {}, 'notas': []}
    current_nota = None

    idx = 0
    while idx < len(lines):
        line = lines[idx].strip()
        if line.startswith('Total de Notas:'):
            notas['resumo']['total_notas'] = int(line.split(':', 1)[1].strip())
        elif line.startswith('Valor Total:'):
            valor_total_str = line.split(':', 1)[1].strip().replace('R$', '').replace(',', '.')
            notas['resumo']['valor_total'] = float(valor_total_str)
        elif line.startswith('Número NFC-e:'):
            if current_nota:
                notas['notas'].append(current_nota)
            current_nota = {}
            current_nota['produtos'] = []
            current_nota['nNF'] = line.split(':', 1)[1].strip()
        elif line.startswith('Nome da Nota:'):
            current_nota['nome'] = line.split(':', 1)[1].strip()
        elif line.startswith('Valor:'):
            valor_str = line.split(':', 1)[1].strip().replace('R$', '').replace(',', '.')
            current_nota['valor'] = float(valor_str)
        elif line.startswith('Status:'):
            current_nota['status'] = line.split(':', 1)[1].strip()
        elif line.startswith('Data de Emissão:'):
            current_nota['emitida'] = line.split(':', 1)[1].strip()
        elif line.startswith('Data de Autorização:'):
            current_nota['autorizada'] = line.split(':', 1)[1].strip()
        elif line.startswith('Produtos:'):
            idx += 1  # Avança para a próxima linha
            while idx < len(lines) and re.match(r'-+\s*Nome:', lines[idx].strip()):
                produto_line = lines[idx].strip()
                produto = {}
                for part in produto_line.split(','):
                    key_value = part.split(':', 1)
                    if len(key_value) == 2:
                        key = key_value[0].strip()
                        key = re.sub(r'^-+\s*', '', key)  # Remove traços e espaços
                        key = remove_accents(key).lower().replace(' ', '_')  # Normaliza a chave
                        value = key_value[1].strip()
                        # Converter valores numéricos
                        if key in ['valor_unitario', 'valor_total']:
                            value = value.replace('R$', '').replace(',', '.')
                            try:
                                value = float(value)
                            except ValueError:
                                value = 0.0
                        # Mantém "quantidade" como string para incluir a unidade
                        produto[key] = value
                # Adiciona o produto somente se tiver dados
                if produto:
                    current_nota['produtos'].append(produto)
                idx += 1
            continue  # Evita incrementar idx novamente
        idx += 1

    if current_nota:
        notas['notas'].append(current_nota)

    return notas

def export_to_pdf(report_text, output_file):
    """
    Exporta o relatório como um arquivo PDF com design aprimorado,
    combinando detalhes da nota fiscal e produtos.
    """
    # [Seu código atual da função export_to_pdf]
    pass  # Substitua por sua implementação atual

def export_to_excel(report_text, output_file):
    """
    Exporta o relatório como um arquivo Excel (.xlsx) formatado.
    """
    try:
        notas = parse_report_text(report_text)

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Aba de Resumo
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"

        ws_resumo['A1'] = "Total de Notas"
        ws_resumo['B1'] = notas['resumo']['total_notas']
        ws_resumo['A2'] = "Valor Total (R$)"
        ws_resumo['B2'] = f"{notas['resumo']['valor_total']:.2f}"

        for cell in ['A1', 'A2']:
            ws_resumo[cell].font = header_font
            ws_resumo[cell].fill = header_fill
            ws_resumo[cell].border = thin_border
            ws_resumo[cell].alignment = center_alignment
        for cell in ['B1', 'B2']:
            ws_resumo[cell].border = thin_border
            ws_resumo[cell].alignment = left_alignment

        # Ajustar largura das colunas na aba Resumo
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 20

        # Criar uma aba para cada Nota Fiscal
        for nota in notas['notas']:
            nfe_number = nota.get('nNF', 'N/A')
            # Limitar o nome da aba a 31 caracteres
            sheet_name = f"NF {nfe_number}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            row = 1

            # Título da Nota Fiscal
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value=f"Nota Fiscal - {nfe_number}")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = center_alignment
            row += 2

            # Detalhes da Nota Fiscal
            details = [
                ['Número NFC-e', nfe_number],
                ['Nome da Nota', nota.get('nome', '')],
                ['Valor (R$)', f"{nota.get('valor', 0):.2f}"],
                ['Status', nota.get('status', '')],
                ['Data de Emissão', nota.get('emitida', '')],
                ['Data de Autorização', nota.get('autorizada', '')],
            ]
            for detail in details:
                ws.cell(row=row, column=1, value=detail[0])
                ws.cell(row=row, column=2, value=detail[1])
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = header_fill
                ws.cell(row=row, column=1).alignment = left_alignment
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2).border = thin_border
                ws.cell(row=row, column=2).alignment = left_alignment
                row += 1

            row += 1  # Espaço entre detalhes e produtos

            # Cabeçalho da Tabela de Produtos
            headers = ['Nome', 'Código', 'CFOP', 'Quantidade', 'Valor Unitário (R$)', 'Valor Total (R$)']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            row += 1

            # Dados dos Produtos
            for prod in nota.get('produtos', []):
                ws.cell(row=row, column=1, value=prod.get('nome', ''))
                ws.cell(row=row, column=2, value=prod.get('codigo', ''))
                ws.cell(row=row, column=3, value=prod.get('cfop', ''))
                ws.cell(row=row, column=4, value=prod.get('quantidade', ''))
                ws.cell(row=row, column=5, value=f"{prod.get('valor_unitario', 0.0):.2f}")
                ws.cell(row=row, column=6, value=f"{prod.get('valor_total', 0.0):.2f}")
                # Formatar células
                for col_num in range(1, 7):
                    cell = ws.cell(row=row, column=col_num)
                    cell.border = thin_border
                    cell.alignment = left_alignment
                row += 1

            # Ajustar largura das colunas
            for col in range(1, 7):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    try:
                        if cell.value:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # Remover a planilha padrão se não for a aba Resumo
        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)

        # Salvar o arquivo Excel
        wb.save(output_file)

    except Exception as e:
        raise RuntimeError(f"Erro ao exportar para Excel: {str(e)}")

def export_to_csv(report_text, output_file):
    """
    Exporta o relatório como um arquivo CSV.
    """
    try:
        notas = parse_report_text(report_text)
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Número NFC-e', 'Nome da Nota', 'Valor (R$)', 'Status',
                          'Data de Emissão', 'Data de Autorização', 'Produtos']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for nota in notas['notas']:
                produtos_str = "; ".join([
                    f"{p.get('nome', '')} (Código: {p.get('codigo', '')}, CFOP: {p.get('cfop', '')}, Quantidade: {p.get('quantidade', '')}, Valor Unitário: {p.get('valor_unitario', '')}, Valor Total: {p.get('valor_total', '')})"
                    for p in nota.get('produtos', []) if p
                ])
                writer.writerow({
                    'Número NFC-e': nota.get('nNF', ''),
                    'Nome da Nota': nota.get('nome', ''),
                    'Valor (R$)': f"{nota.get('valor', 0):.2f}",
                    'Status': nota.get('status', ''),
                    'Data de Emissão': nota.get('emitida', ''),
                    'Data de Autorização': nota.get('autorizada', ''),
                    'Produtos': produtos_str
                })
    except Exception as e:
        raise RuntimeError(f"Erro ao exportar para CSV: {str(e)}")

def export_to_txt(report_text, output_file):
    """
    Exporta o relatório como um arquivo TXT.
    """
    try:
        with open(output_file, 'w', encoding='utf-8') as txtfile:
            txtfile.write(report_text)
    except Exception as e:
        raise RuntimeError(f"Erro ao exportar para TXT: {str(e)}")
